from tkinter import *
from tkinter.ttk import Combobox, Treeview, Style
from tkinter import font
from tkinter import messagebox
from datetime import date, datetime, timedelta
import time
from openpyxl import Workbook, load_workbook
import os
import matplotlib.pyplot as plt
from collections import defaultdict

root = Tk()
root.title("Muhasebe")
root.geometry('900x500')
root.maxsize(900, 500)
root.minsize(900, 500)

# Yazı tipi ayarları
label_font = font.Font(family='Helvetica', size=16, weight='bold')

def kurallar():
    messagebox.showinfo("kurallar","her dosya ayın 15 inden itibaren tutulmaya başlar.dosyalar masaüstündeki klasörün içinde tutulur.!!!klasörün yerini değiştirmeyiniz masaüstünde kalsın.  girdi kategorisine girdi yazınız.grafik sadece çıktıları gösterir.tarih değiştirilemez.girdi çiktı butonlarının seçili oldupundan emin olunuz(default çıktıdır).")

def get_desktop_directory():
    # Masaüstü dizinine gitmek için kullanıcı yolunu al
    if os.name == 'nt':  # Windows için
        desktop = os.path.join(os.environ["HOMEPATH"], 'Desktop')
    else:  # Mac ve Linux için
        desktop = os.path.join(os.path.expanduser("~"), 'Desktop')
    return desktop

def get_days_until_next_month_15():
    today = date.today()

    # Eğer bugün ayın 15'ine kadar olan bir tarihe denk geliyorsa, bu ayın 15'ine kadar kalan gün sayısını hesapla
    if today.day <= 15:
        next_15 = today.replace(day=15)  # Bu ayın 15'i
    else:
        # Eğer bugün ayın 15'inden sonra bir tarihse, bir sonraki ayın 15'ine kadar kalan gün sayısını hesapla
        next_month = today.replace(month=today.month % 12 + 1, day=15)  # Bir sonraki ayın 15'i
        next_15 = next_month

    # Kalan gün sayısını hesapla
    days_until_15 = (next_15 - today).days
    return days_until_15

# Dosya ismini yıl ve ay bilgisini alacak şekilde oluştur
def get_monthly_filename():
    # Masaüstünde "Muhasebe Kayıtları" klasörünü kontrol et, yoksa oluştur
    desktop_dir = get_desktop_directory()
    muhasebe_dir = os.path.join(desktop_dir, "Muhasebe Kayıtları")
    
    # Klasör yoksa oluştur
    if not os.path.exists(muhasebe_dir):
        os.makedirs(muhasebe_dir)
    
    # Excel dosyasının tam yolu
    current_date = datetime.now()
    file_name = os.path.join(muhasebe_dir, f"muhasebe_{current_date.year}_{current_date.month:02d}.xlsx")
    
    return file_name

def plot_graph():
    try:
        file_name = get_monthly_filename()
        # Excel dosyasını yükle
        workbook = load_workbook(file_name)
        sheet = workbook.active

        # Harcama verilerini depolamak için bir dictionary (kategori: toplam tutar)
        category_data = defaultdict(float)

        # Excel dosyasındaki tüm satırlarda dolaşarak "çıktı" harcamalarını topla
        for row in sheet.iter_rows(min_row=2, values_only=True):
            kategori_value = row[1]  # Kategori
            tutar = row[3]           # Tutar
            durum = row[4]           # Durum ("girdi" veya "çıktı")

            # Sadece "çıktı" olan verileri al
            if durum == "cıktı":
                category_data[kategori_value] += tutar

        # Eğer harcama verisi varsa
        if category_data:
            # Grafik verisini hazırla
            categories = list(category_data.keys())
            amounts = list(category_data.values())

            # Pasta grafiğini çiz
            plt.figure(figsize=(8, 8))
            plt.pie(amounts, labels=categories, autopct='%1.1f%%', startangle=140)
            plt.title('Çıktı Harcama Dağılımı')
            plt.axis('equal')  # Yatay dikey oranı eşitle, yuvarlak pasta grafik elde et
            plt.show()
            # Harcama detaylarını göstermek için metin
            total_spent = sum(amounts)
            result_text = "Toplam Harcama: {:.2f} TL\n\n".format(total_spent)
            for category, amount in category_data.items():
                result_text += f"{category}: {amount:.2f} TL ({(amount / total_spent) * 100:.1f}%)\n"
           
           
            messagebox.showinfo("Harcama Detayları", result_text)
        else:
            messagebox.showinfo("Veri Yok", "Çıktı harcaması bulunamadı.")
    except FileNotFoundError:
        messagebox.showerror("Hata", "Excel dosyası bulunamadı.")

def girdicikti():
    return "girdi" if durumgirdicikti.get() == '1' else "cıktı"

def kaydet():
    kategori_value = kategori.get()
    aciklama_value = aciklamaentry.get()
    fiyat_value = float(fiyatentry.get())
    durum_value = girdicikti()
    
    # Kayıt bilgilerini göster ve kullanıcıdan onay al
    onay_mesaji = f"Tarih: {tarih}\nKategori: {kategori_value}\nAçıklama: {aciklama_value}\nTutar: {fiyat_value}\nDurum: {durum_value}\n\nKayıt yapılsın mı?"
    if not messagebox.askyesno("Onay", onay_mesaji):
        return

    # Dosya ismini al
    file_name = get_monthly_filename()

    # Load or create the Excel workbook
    try:
        workbook = load_workbook(file_name)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Tarih", "Kategori", "Açıklama", "Tutar", "Durum"])
        # İlk kayıt tarihi
        sheet['H1'] = date.today().strftime("%d/%m/%Y")
    
    # Append new data to the workbook
    sheet.append([tarih, kategori_value, aciklama_value, fiyat_value, durum_value])
    workbook.save(file_name)
    
    # Update balances and records
    update_balances()
    update_records()

def update_balances():
    try:
        file_name = get_monthly_filename()
        workbook = load_workbook(file_name)
        sheet = workbook.active

        total_income = 0
        total_expense = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[4] == "girdi":
                total_income += row[3]
            else:
                total_expense += row[3]

        kalan_para = total_income - total_expense
        harcanan_para = total_expense

        kalan_para_label.config(text=f"Kalan Para: {kalan_para} TL")
        harcanan_para_label.config(text=f"Harcanan Para: {harcanan_para} TL")

        # Sonraki ayın 15'ine kalan gün sayısını hesapla
        remaining_days = get_days_until_next_month_15()
        gecen_gun_label.config(text=f"Sonraki Ayın 15'ine Kalan Gün: {remaining_days} gün")

    except FileNotFoundError:
        kalan_para_label.config(text="Kalan Para: 0 TL")
        harcanan_para_label.config(text="Harcanan Para: 0 TL")
        gecen_gun_label.config(text="Sonraki Ayın 15'ine Kalan Gün: 0 gün")

def update_records():
    try:
        for i in treeview.get_children():
            treeview.delete(i)
        file_name = get_monthly_filename()
        workbook = load_workbook(file_name)
        sheet = workbook.active

        records = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            records.insert(0, row)  # Yeni eklenen kaydı listenin başına ekle
        
        for record in records:
            treeview.insert('', 'end', values=record)
    
    except FileNotFoundError:
        pass

def sil():
    # Seçilen öğeyi al
    selected_item = treeview.selection()[0]  # Seçilen öğe
    values = treeview.item(selected_item, 'values')

    # Silme onayı al
    onay_mesaji = f"Tarih: {values[0]}\nKategori: {values[1]}\nAçıklama: {values[2]}\nTutar: {values[3]}\nDurum: {values[4]}\n\nKaydı silmek istediğinizden emin misiniz?"
    if not messagebox.askyesno("Silme Onayı", onay_mesaji):
        return  # Kullanıcı silme işlemine onay vermezse işlemi iptal et

    # Excel'den silme işlemi
    try:
        # Dosya ismini al
        file_name = get_monthly_filename()

        # Excel dosyasını aç
        workbook = load_workbook(file_name)
        sheet = workbook.active

        # Excel dosyasındaki tüm satırları kontrol et
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            # Satırdaki tüm hücrelerin eşleştiğinden emin ol
            if (row[0].value == values[0] and 
                row[1].value == values[1] and 
                row[2].value == values[2] and 
                row[3].value == float(values[3]) and  # Fiyatı float olarak karşılaştır
                row[4].value == values[4]):
                sheet.delete_rows(row[0].row, 1)  # Satırı sil
                break  # İlk eşleşmeyi bulunca döngüyü bitir

        workbook.save(file_name)  # Dosyayı kaydet
    except FileNotFoundError:
        messagebox.showerror("Hata", "Excel dosyası bulunamadı.")  # Dosya bulunamazsa hata mesajı

    # Treeview'den silme işlemi
    treeview.delete(selected_item)

    # Bakiye ve diğer bilgileri güncelle
    update_balances()  # Güncel bakiye bilgilerini yeniden hesapla ve göster
    update_records()  # Treeview'i güncelle
    
def guncelle():
    selected_item = treeview.selection()[0]  # Seçilen öğe
    values = treeview.item(selected_item, 'values')

    kategori.set(values[1])
    aciklamaentry.delete(0, END)
    aciklamaentry.insert(0, values[2])
    fiyatentry.delete(0, END)
    fiyatentry.insert(0, values[3])

    def on_save_update():
        kategori_value = kategori.get()
        aciklama_value = aciklamaentry.get()
        fiyat_value = float(fiyatentry.get())
        durum_value = girdicikti()

        # Kayıt bilgilerini göster ve kullanıcıdan onay al
        onay_mesaji = f"Tarih: {tarih}\nKategori: {kategori_value}\nAçıklama: {aciklama_value}\nTutar: {fiyat_value}\nDurum: {durum_value}\n\nKayıt güncellensin mi?"
        if not messagebox.askyesno("Onay", onay_mesaji):
            return

        # Excel dosyasını aç
        file_name = get_monthly_filename()

        try:
            workbook = load_workbook(file_name)
            sheet = workbook.active

            # Seçilen öğeyi bul ve güncelle
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                if (row[0].value == values[0] and 
                    row[1].value == values[1] and 
                    row[2].value == values[2] and 
                    row[3].value == float(values[3]) and 
                    row[4].value == values[4]):
                    row[1].value = kategori_value
                    row[2].value = aciklama_value
                    row[3].value = fiyat_value
                    row[4].value = durum_value
                    break

            workbook.save(file_name)
        except FileNotFoundError:
            messagebox.showerror("Hata", "Excel dosyası bulunamadı.")

        # Treeview'deki veriyi güncelle
        treeview.item(selected_item, values=(tarih, kategori_value, aciklama_value, fiyat_value, durum_value))

        # Bakiye ve diğer bilgileri güncelle
        update_balances()
        update_records()
        save_button.place_forget()
    save_button = Button(root, text="Kaydet ve Güncelle", command=on_save_update,)
    save_button.place(x=480, y=350, width=100, height=50)
# tarih
tarih = date.today()
tarih = tarih.strftime("%d/%m/%Y")
tarihlabel = Label(root, text=tarih, bg='pink')
tarihlabel.place(x=740, y=10, width=150, height=50)

# kategori
kategorilabel = Label(root, text="kategori:")
kategorilabel.place(x=10, y=10)

kategori = Combobox(root, values=('yemek', 'eğlence', 'barınma', 'ulaşım', 'hobi', 'giyim', 'sağlık', 'eğitim', 'sigara', 'girdi','borç','diğer'))
kategori.place(x=70, y=10)

# açıklama
aciklamaLabel1 = Label(root, text="açıklama:")
aciklamaLabel1.place(x=10, y=30)

aciklamaentry = Entry()
aciklamaentry.place(x=70, y=30) 

# fiyat
fiyatLabel = Label(root, text="tutar:")
fiyatLabel.place(x=10, y=50)

fiyatentry = Entry()
fiyatentry.place(x=70, y=50) 

# tarih giriş alanı
tarihentry = Entry()
tarihentry.place(x=250, y=10)

# radiobutton
durumgirdicikti = StringVar()

girdi = Radiobutton(root, text="girdi", bg='green', value=1, variable=durumgirdicikti)
girdi.place(x=10, y=70, width=55)

cikti = Radiobutton(root, text="cıktı", bg='red', value=2, variable=durumgirdicikti)
cikti.place(x=10, y=95, width=55)

# ekle buton
eklebuton = Button(root, text="kaydı ekle", bg='gray', command=kaydet)
eklebuton.place(x=10, y=130, width=100, height=50)

# sil butonu
silbuton = Button(root, text="kaydı sil", bg='gray', command=sil)
silbuton.place(x=320, y=400, width=100, height=50)

# güncelle butonu
guncellebuton = Button(root, text="güncelle", bg='gray', command=guncelle)
guncellebuton.place(x=480, y=400, width=100, height=50)
# Create a button to generate the category graph

#grafik buton
grafik_buton = Button(root, text="Kategorilere Göre Grafik Oluştur", bg='blue', command=plot_graph)
grafik_buton.place(x=10, y=200, width=100, height=50)

#button kurallar
kuralbuton = Button(root,text="nasıl çalışır",bg='pink',command=kurallar)
kuralbuton.place(x=720, y=150, width=150, height=50)

# kalan para ve harcanan para etiketleri
kalan_para_label = Label(root, text="Kalan Para: 0 TL", bg='green')
kalan_para_label.place(x=720, y=250, width=150, height=50)

harcanan_para_label = Label(root, text="Harcanan Para: 0 TL", bg='red')
harcanan_para_label.place(x=720, y=300, width=150, height=50)

# ilk kayıttan sonra geçen gün sayısı etiketi
gecen_gun_label = Label(root, text="Geçen Gün: 0 gün",bg='pink')
gecen_gun_label.place(x=700, y=70, width=200, height=50)

# Treeview widget'ı
style = Style()
style.configure('Treeview', rowheight=25)
treeview = Treeview(root, columns=('Tarih', 'Kategori', 'Açıklama', 'Tutar', 'Durum'), show='headings')
treeview.heading('Tarih', text='Tarih')
treeview.heading('Kategori', text='Kategori')
treeview.heading('Açıklama', text='Açıklama')
treeview.heading('Tutar', text='Tutar')
treeview.heading('Durum', text='Durum')

# Sütun genişliklerini ayarla
treeview.column('Tarih', width=80)
treeview.column('Kategori', width=80)
treeview.column('Açıklama', width=120)
treeview.column('Tutar', width=80)
treeview.column('Durum', width=80)

treeview.place(x=250, y=70, width=450, height=180)

# Uygulama açıldığında verileri yükle
update_balances()
update_records()

root.mainloop()